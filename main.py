import os
import json
import datetime
import urllib.parse
import http.client
import requests
from flask import Flask, request
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font
from io import BytesIO

# ===================== CONFIG =====================
API_KEY = os.getenv("API_KEY")  # Football API key
TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN")
CHAT_ID = os.getenv("CHAT_ID")  # Bisa fixed ID atau ambil dari update
PORT = int(os.getenv("PORT", 5000))  # Railway pakai PORT ini

headers = {
    "x-rapidapi-host": "v3.football.api-sports.io",
    "x-rapidapi-key": API_KEY
}

TIMEZONE = "Asia/Makassar"
ALLOWED_LEAGUES = [39, 135, 140, 703]  # EPL, Serie A, La Liga
# ==================================================

app = Flask(__name__)

# ===================== API FUNCTIONS =====================
def get_fixtures_today_and_tomorrow():
    fixtures = []
    tz_encoded = urllib.parse.quote(TIMEZONE)

    for offset in [0, 1]:  # 0 = hari ini, 1 = besok
        target_date = (datetime.date.today() + datetime.timedelta(days=offset)).strftime("%Y-%m-%d")
        params = f"/fixtures?date={target_date}&status=NS&timezone={tz_encoded}"

        conn = http.client.HTTPSConnection("v3.football.api-sports.io")
        conn.request("GET", params, headers=headers)
        res = conn.getresponse()
        data = res.read()
        conn.close()

        json_data = json.loads(data.decode("utf-8"))
        fixtures.extend(json_data.get("response", []))

    return fixtures


def save_filtered_fixtures(filename="fixtures.json"):
    fixtures = get_fixtures_today_and_tomorrow()

    result = []
    for match in fixtures:
        fixture_id = match["fixture"]["id"]
        league_id = match["league"]["id"]

        if league_id in ALLOWED_LEAGUES:
            result.append({
                "fixture_id": fixture_id,
                "league_id": league_id
            })

    with open(filename, "w", encoding="utf-8") as f:
        json.dump(result, f, indent=2)


def get_prediction(fixture_id):
    conn = http.client.HTTPSConnection("v3.football.api-sports.io")
    params = f"/predictions?fixture={fixture_id}"
    conn.request("GET", params, headers=headers)
    res = conn.getresponse()
    data = res.read()
    conn.close()
    return json.loads(data.decode("utf-8"))


def load_fixture_ids(filename="fixtures.json"):
    with open(filename, "r", encoding="utf-8") as f:
        return json.load(f)


def save_predictions(fixtures_file="fixtures.json", predictions_file="predictions.json"):
    fixtures = load_fixture_ids(fixtures_file)
    results = []

    for match in fixtures:
        fixture_id = match["fixture_id"]
        data = get_prediction(fixture_id)
        if not data.get("response"):
            continue

        resp = data["response"][0]
        prediction = {
            "fixture_id": fixture_id,
            "league": resp["league"]["name"],
            "advice": resp["predictions"]["advice"],
            "home_last5": resp["teams"]["home"]["last_5"],
            "away_last5": resp["teams"]["away"]["last_5"],
            "home_form": resp["teams"]["home"]["league"]["form"],
            "away_form": resp["teams"]["away"]["league"]["form"],
            "comparison": {
                "h2h": resp["comparison"]["h2h"],
                "total": resp["comparison"]["total"]
            }
        }
        results.append(prediction)

    with open(predictions_file, "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2, ensure_ascii=False)


def load_predictions(filename="predictions.json"):
    with open(filename, "r", encoding="utf-8") as f:
        return json.load(f)

# ===================== EXCEL BUILDER =====================
def build_predictions_excel(predictions):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Predictions"

    headers_main = [
        "Fixture ID", "League", "Advice",
        "Last 5 Played", "", "Form", "", "Att", "", "Def", "",
        "League Form", "", "H2H", "", "Total", ""
    ]
    ws.append(headers_main)

    headers_sub = [
        "", "", "",
        "Home", "Away",
        "Home", "Away",
        "Home", "Away",
        "Home", "Away",
        "Home", "Away",
        "Home", "Away",
        "Home", "Away"
    ]
    ws.append(headers_sub)

    merge_ranges = [
        "A1:A2", "B1:B2", "C1:C2",
        "D1:E1", "F1:G1", "H1:I1", "J1:K1",
        "L1:M1", "N1:O1", "P1:Q1"
    ]
    for rng in merge_ranges:
        ws.merge_cells(rng)

    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    bold_font = Font(bold=True)

    for row in ws.iter_rows(min_row=1, max_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = bold_font
            cell.fill = header_fill

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for p in predictions:
        row = [
            p["fixture_id"],
            p["league"],
            p["advice"],
            p["home_last5"].get("played", ""),
            p["away_last5"].get("played", ""),
            p["home_last5"].get("form", ""),
            p["away_last5"].get("form", ""),
            p["home_last5"].get("att", ""),
            p["away_last5"].get("att", ""),
            p["home_last5"].get("def", ""),
            p["away_last5"].get("def", ""),
            p["home_form"],
            p["away_form"],
            p["comparison"]["h2h"]["home"],
            p["comparison"]["h2h"]["away"],
            p["comparison"]["total"]["home"],
            p["comparison"]["total"]["away"],
        ]
        ws.append(row)

        row_idx = ws.max_row
        compare_pairs = [(4, 5), (6, 7), (8, 9), (10, 11), (14, 15), (16, 17)]

        for home_col, away_col in compare_pairs:
            home_val = ws.cell(row=row_idx, column=home_col).value
            away_val = ws.cell(row=row_idx, column=away_col).value
            try:
                h = float(str(home_val).replace("%", "").strip())
                a = float(str(away_val).replace("%", "").strip())
            except:
                continue

            if h > a:
                ws.cell(row=row_idx, column=home_col).fill = green_fill
                ws.cell(row=row_idx, column=away_col).fill = red_fill
            elif h < a:
                ws.cell(row=row_idx, column=home_col).fill = red_fill
                ws.cell(row=row_idx, column=away_col).fill = green_fill

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value and len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        ws.column_dimensions[col_letter].width = max_length + 2

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream

# ===================== TELEGRAM FUNCTIONS =====================
def send_message(chat_id, text):
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
    payload = {"chat_id": chat_id, "text": text, "parse_mode": "HTML"}
    requests.post(url, json=payload)

def send_predictions_excel(chat_id, predictions):
    excel_file = build_predictions_excel(predictions)
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendDocument"
    files = {"document": ("predictions.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    data = {"chat_id": chat_id}
    requests.post(url, data=data, files=files)

# ===================== WEBHOOK HANDLER =====================
@app.route(f"/{TELEGRAM_TOKEN}", methods=["POST"])
def telegram_webhook():
    update = request.get_json()
    if not update or "message" not in update:
        return "ok"

    chat_id = update["message"]["chat"]["id"]
    text = update["message"].get("text", "")

    if text == "/start":
        send_message(chat_id, "👋 Halo! Ketik /prediksi untuk melihat prediksi (hari ini + besok) dalam bentuk file Excel.")
    elif text == "/prediksi":
        save_filtered_fixtures()
        save_predictions()
        preds = load_predictions()

        if not preds:
            send_message(chat_id, "⚽ Tidak ada prediksi untuk hari ini & besok.")
        else:
            send_predictions_excel(chat_id, preds)

    return "ok"

# ===================== MAIN =====================
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=PORT)
