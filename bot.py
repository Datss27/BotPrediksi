import http.client
import os
import json
import datetime
import urllib.parse
import openpyxl
from openpyxl.styles import Alignment, PatternFill
from telegram import Update
from telegram.ext import Application, CommandHandler, ContextTypes
import sys

# ===================== CONFIG =====================
API_KEY = os.getenv("API_KEY")  # Football API key
TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN")
CHAT_ID = 7952198349

headers = {
    "x-rapidapi-host": "v3.football.api-sports.io",
    "x-rapidapi-key": API_KEY
}

TIMEZONE = "Asia/Makassar"
ALLOWED_LEAGUES = [39, 135, 140, 703]  # EPL, Serie A, La Liga
# ==================================================

# ===================== API FUNCTIONS =====================
def get_today_fixtures():
    today = datetime.date.today().strftime("%Y-%m-%d")
    tz_encoded = urllib.parse.quote(TIMEZONE)
    params = f"/fixtures?date={today}&status=NS&timezone={tz_encoded}"

    conn = http.client.HTTPSConnection("v3.football.api-sports.io")
    conn.request("GET", params, headers=headers)
    res = conn.getresponse()
    data = res.read()
    conn.close()
    return json.loads(data.decode("utf-8"))

def save_filtered_fixtures(filename="fixtures.json"):
    data = get_today_fixtures()
    fixtures = data.get("response", [])

    result = []
    for match in fixtures:
        fixture_id = match["fixture"]["id"]
        league_id = match["league"]["id"]

        if league_id in ALLOWED_LEAGUES:
            result.append({
                "fixture_id": fixture_id,
                "league_id": league_id
            })

    with open(filename, "w", encoding="utf-8") as f:
        json.dump(result, f, indent=2)

def get_prediction(fixture_id):
    conn = http.client.HTTPSConnection("v3.football.api-sports.io")
    params = f"/predictions?fixture={fixture_id}"
    conn.request("GET", params, headers=headers)
    res = conn.getresponse()
    data = res.read()
    conn.close()
    return json.loads(data.decode("utf-8"))

def load_fixture_ids(filename="fixtures.json"):
    with open(filename, "r", encoding="utf-8") as f:
        return json.load(f)

def save_predictions(fixtures_file="fixtures.json", predictions_file="predictions.json"):
    fixtures = load_fixture_ids(fixtures_file)
    results = []

    for match in fixtures:
        fixture_id = match["fixture_id"]
        data = get_prediction(fixture_id)
        if not data.get("response"):
            continue

        resp = data["response"][0]
        prediction = {
            "fixture_id": fixture_id,
            "league": resp["league"]["name"],
            "advice": resp["predictions"]["advice"],
            "home_last5": resp["teams"]["home"]["last_5"],
            "away_last5": resp["teams"]["away"]["last_5"],
            "home_form": resp["teams"]["home"]["league"]["form"],
            "away_form": resp["teams"]["away"]["league"]["form"],
            "comparison": {
                "h2h": resp["comparison"]["h2h"],
                "total": resp["comparison"]["total"]
            }
        }
        results.append(prediction)

    with open(predictions_file, "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2, ensure_ascii=False)

def load_predictions(filename="predictions.json"):
    with open(filename, "r", encoding="utf-8") as f:
        return json.load(f)

def save_predictions_to_excel(predictions, excel_file="predictions.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Predictions"

    headers_main = [
        "Fixture ID", "League", "Advice",
        "Last 5 Played", "", "Form", "", "Att", "", "Def", "",
        "League Form", "", "H2H", "", "Total", ""
    ]
    ws.append(headers_main)

    headers_sub = [
        "", "", "",
        "Home", "Away",
        "Home", "Away",
        "Home", "Away",
        "Home", "Away",
        "Home", "Away",
        "Home", "Away",
        "Home", "Away"
    ]
    ws.append(headers_sub)

    ws.merge_cells("A1:A2")
    ws.merge_cells("B1:B2")
    ws.merge_cells("C1:C2")
    ws.merge_cells("D1:E1")
    ws.merge_cells("F1:G1")
    ws.merge_cells("H1:I1")
    ws.merge_cells("J1:K1")
    ws.merge_cells("L1:M1")
    ws.merge_cells("N1:O1")
    ws.merge_cells("P1:Q1")

    for row in ws.iter_rows(min_row=1, max_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for p in predictions:
        row = [
            p["fixture_id"],
            p["league"],
            p["advice"],
            p["home_last5"].get("played", ""),
            p["away_last5"].get("played", ""),
            p["home_last5"].get("form", ""),
            p["away_last5"].get("form", ""),
            p["home_last5"].get("att", ""),
            p["away_last5"].get("att", ""),
            p["home_last5"].get("def", ""),
            p["away_last5"].get("def", ""),
            p["home_form"],
            p["away_form"],
            p["comparison"]["h2h"]["home"],
            p["comparison"]["h2h"]["away"],
            p["comparison"]["total"]["home"],
            p["comparison"]["total"]["away"],
        ]
        ws.append(row)

        row_idx = ws.max_row
        compare_pairs = [(4, 5), (6, 7), (8, 9), (10, 11), (12, 13), (14, 15), (16, 17)]

        for home_col, away_col in compare_pairs:
            home_val = ws.cell(row=row_idx, column=home_col).value
            away_val = ws.cell(row=row_idx, column=away_col).value
            try:
                h = float(str(home_val).replace("%", "").strip())
                a = float(str(away_val).replace("%", "").strip())
            except:
                continue

            if h > a:
                ws.cell(row=row_idx, column=home_col).fill = green_fill
                ws.cell(row=row_idx, column=away_col).fill = red_fill
            elif h < a:
                ws.cell(row=row_idx, column=home_col).fill = red_fill
                ws.cell(row=row_idx, column=away_col).fill = green_fill

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value and len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(excel_file)
# ==================================================

# ===================== TELEGRAM BOT =====================
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("⚽ Halo! Kirim /prediksi untuk dapatkan prediksi hari ini.")

async def prediksi(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("🔄 Mengambil data...")

    save_filtered_fixtures()
    save_predictions()

    predictions = load_predictions()
    excel_file = "predictions.xlsx"
    save_predictions_to_excel(predictions, excel_file)

    await update.message.reply_document(document=open(excel_file, "rb"), filename=excel_file)
    # Hapus file setelah dikirim
    for f in ["fixtures.json", "predictions.json", "predictions.xlsx"]:
        if os.path.exists(f):
            os.remove(f)
# ==================================================

def main():
    app = Application.builder().token(TELEGRAM_TOKEN).build()

    if "--auto" in sys.argv:
        # mode otomatis
        save_filtered_fixtures()
        save_predictions()
        predictions = load_predictions()
        excel_file = "predictions.xlsx"
        save_predictions_to_excel(predictions, excel_file)

        async def send_file():
            await app.bot.send_document(
                chat_id=CHAT_ID,
                document=open(excel_file, "rb")
            )
            # hapus file
            for f in ["fixtures.json", "predictions.json", "predictions.xlsx"]:
                if os.path.exists(f):
                    os.remove(f)

        # jalankan hanya sekali
        asyncio.run(send_file())

    else:
        # mode normal (polling)
        app.add_handler(CommandHandler("start", start))
        app.add_handler(CommandHandler("prediksi", prediksi))
        app.run_polling()

if __name__ == "__main__":
    main()
