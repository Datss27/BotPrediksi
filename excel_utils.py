from io import BytesIO
from datetime import datetime
from zoneinfo import ZoneInfo
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from settings import settings

TZ = ZoneInfo(settings.timezone)

COLOR_GREEN = "FF63BE7B"
COLOR_RED = "FFF8696B"
COLOR_YELLOW = "FFADD8E6"
WHITE_RGB = "FFFFFF"

def _parse_percent(value):
    if isinstance(value, str) and "%" in value:
        try:
            return float(value.replace("%", ""))
        except ValueError:
            return "-"
    return value if isinstance(value, (int, float)) else "-"

def blend_color(hex_color, factor):
    """Campur warna berdasarkan faktor jarak."""
    r = int(hex_color[0:2], 16)
    g = int(hex_color[2:4], 16)
    b = int(hex_color[4:6], 16)
    white = 255
    new_r = int((1 - factor) * white + factor * r)
    new_g = int((1 - factor) * white + factor * g)
    new_b = int((1 - factor) * white + factor * b)
    return f"{new_r:02X}{new_g:02X}{new_b:02X}"

def safe_int(value):
    try:
        return int(value)
    except:
        return 0

def calculate_performance(wins, draws, played):
    if played == 0:
        return 0.0  # Hindari divide by zero
    max_points = played * 3
    actual_points = (wins * 3) + (draws * 1)
    return round(actual_points / max_points, 3)

def create_workbook(fixtures):
    wb = Workbook()
    ws = wb.active
    date_str = datetime.now(TZ).strftime("%Y-%m-%d")
    ws.title = f"Prediksi {date_str}"

    headers = [
        "Negara", "Liga", "Home", "Away", "Tanggal", "Jam", "Saran",
        "Probabilitas (H/D/A)",
        "History", None,
        "Form", None,
        "ATT", None,
        "DEF", None,
        "Comp", None,
        "H2H", None
    ]
    subheaders = [""] * 8 + ["Home", "Away"] * 6

    ws.append(headers)
    ws.append(subheaders)

    # Style header
    header_fill = PatternFill("solid", fgColor="FFFF00")
    for row in ws.iter_rows(min_row=1, max_row=2):
        for cell in row:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = header_fill

    for col in range(1, 9):
        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)

    merge_groups = {
        "Performance": (9, 10),
        "Form": (11, 12),
        "ATT": (13, 14),
        "DEF": (15, 16),
        "Comp": (17, 18),
        "H2H": (19, 20)
    }
    for _, (start_col, end_col) in merge_groups.items():
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

    written_matches = 0
    for f in fixtures:
        row = _extract_row(f)
        if not row:
            continue
        ws.append(row)
        written_matches += 1

    # Terapkan gradasi warna ke kolom perbandingan
    for row_idx in range(3, ws.max_row + 1):
        for (h_col, a_col) in [(11, 12), (13, 14), (15, 16), (17, 18), (19, 20)]:
            h_cell = ws.cell(row=row_idx, column=h_col)
            a_cell = ws.cell(row=row_idx, column=a_col)
            try:
                hv = float(h_cell.value)
                av = float(a_cell.value)
            except (ValueError, TypeError):
                continue

            max_val = max(hv, av)
            diff = abs(hv - av)
            factor = diff / max_val if max_val != 0 else 0

            if hv > av:
                color_h = blend_color(COLOR_GREEN[2:], factor)
                color_a = f"FF{WHITE_RGB}"
            elif av > hv:
                color_a = blend_color(COLOR_RED[2:], factor)
                color_h = f"FF{WHITE_RGB}"
            else:
                color_h = color_a = COLOR_YELLOW

            h_cell.fill = PatternFill("solid", fgColor=color_h)
            a_cell.fill = PatternFill("solid", fgColor=color_a)

    for i, col_cells in enumerate(ws.columns, 1):
        col_letter = get_column_letter(i)
        max_len = max((len(str(c.value)) for c in col_cells if c.value), default=0)
        ws.column_dimensions[col_letter].width = max_len + 2

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio, written_matches

def _extract_row(f):
    pred = f.get('prediction') or []
    if not pred or not pred[0].get('predictions'):
        return None

    p = pred[0]
    pr = p.get('predictions', {})
    advice = pr.get('advice', '-')
    pct = pr.get('percent', {})
    hp = pct.get('home', '-')
    dp = pct.get('draw', '-')
    ap = pct.get('away', '-')
    prob_summary = f"{hp} / {dp} / {ap}"
    t = p.get('teams', {})
    home, away = t.get('home', {}), t.get('away', {})
    # Ambil statistik home dan away
    home_stats = home.get("league", {}).get("fixtures", {})
    away_stats = away.get("league", {}).get("fixtures", {})

    # Home team
    h_played = safe_int(home_stats.get("played", {}).get("total", "-"))
    h_wins = safe_int(home_stats.get("wins", {}).get("total", "-"))
    h_draws = safe_int(home_stats.get("draws", {}).get("total", "-"))
    h_loses = safe_int(home_stats.get("loses", {}).get("total", "-"))

    # Away team
    a_played = safe_int(away_stats.get("played", {}).get("total", "-"))
    a_wins = safe_int(away_stats.get("wins", {}).get("total", "-"))
    a_draws = safe_int(away_stats.get("draws", {}).get("total", "-"))
    a_loses = safe_int(away_stats.get("loses", {}).get("total", "-"))
    home_sum = calculate_performance(h_wins, h_draws, h_played)
    away_sum = calculate_performance(a_wins, a_draws, a_played)
    form = _parse_percent(home.get('last_5', {}).get('form'))
    form_away = _parse_percent(away.get('last_5', {}).get('form'))
    att = _parse_percent(home.get('last_5', {}).get('att'))
    att_away = _parse_percent(away.get('last_5', {}).get('att'))
    df = _parse_percent(home.get('last_5', {}).get('def'))
    df_away = _parse_percent(away.get('last_5', {}).get('def'))

    comparison = p.get("comparison", {}).get("total", {})
    comp_home = _parse_percent(comparison.get("home"))
    comp_away = _parse_percent(comparison.get("away"))
    hth = p.get("comparison", {}).get("h2h", {})
    hth_home = _parse_percent(hth.get("home"))
    hth_away = _parse_percent(hth.get("away"))

    dt = datetime.fromisoformat(f['fixture']['date'].replace('Z', '+00:00')).astimezone(TZ)
    date = dt.strftime("%d-%m-%Y")
    time = dt.strftime("%H:%M %Z")

    return [
        f['league']['country'],
        f['league']['name'],
        f['teams']['home']['name'],
        f['teams']['away']['name'],
        date,
        time,
        advice,
        prob_summary,
        home_sum, away_sum,
        form, form_away,
        att, att_away,
        df, df_away,
        comp_home, comp_away,
        hth_home, hth_away
    ]
